import streamlit as st
import os
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
import io
import pandas as pd
import numpy as np

# Define styling constant
DATA_NEED_ATTENTION = 'background-color: #FF0000; color: white; font-weight: bold;'
HIGHLIGHT_NA_ROWS = 'background-color: #FFC7CE;' # Light red for all N/A rows

# Set Streamlit page configuration
st.set_page_config(
    page_title="Sonar Metrics Dashboard",
    layout="wide",
    initial_sidebar_state="auto",
    page_icon=":bar_chart:",
)

st.title("Sonar Metrics Dashboard")


# st.markdown("This tool fetches SonarCloud metrics for your projects and exports them to an Excel file.")

# Global variables for authentication and base URL
auth = None
SONAR_BASE_URL = "" # This will be set from UI input
ORG_NAME = "" # This will be set from UI input
ORG_KEY = "" # This will be set from UI input

# Mapping for numeric ratings to letter grades
RATING_MAP = {
    "1.0": "A",
    "2.0": "B",
    "3.0": "C",
    "4.0": "D",
    "5.0": "E",
    "N/A": "N/A"
}

# ----------------------------- Helper Functions -----------------------------

def make_request(url_path, params=None):
    """
    Makes a GET request to the given URL path, using the global SONAR_BASE_URL.
    """
    global SONAR_BASE_URL # Ensure we're using the global variable
    full_url = f"{SONAR_BASE_URL}{url_path}"
    try:
        response = requests.get(full_url, auth=auth, params=params)
        response.raise_for_status()

        if response.status_code != 200:
            st.error(f"Error: {response.status_code} - {response.text}")
            return None
        if response.status_code == 204:
            st.warning("No content returned.")
            return None
        return response.json()
    except requests.exceptions.RequestException as e:
        st.error(f"Request Error: {e}")
        return None

def fetch_projects(organization_key):
    """
    Fetches all projects for the given organization key.
    """
    url_path = "/api/components/search"
    params = {"organization": organization_key}
    data = make_request(url_path, params)
    return data.get("components", []) if data else []

def fetch_project_metrics(project_key, metric_keys):
    """
    Fetches metrics for a specific project.
    """
    url_path = "/api/measures/component"
    params = {"component": project_key, "metricKeys": metric_keys}
    data = make_request(url_path, params)
    return data.get("component", {}).get("measures", []) if data else []

def fetch_last_analysis_date(project_key):
    """
    Fetches the last analysis date for a specific project.
    """
    url_path = "/api/project_analyses/search"
    params = {"project": project_key, "ps": 1}
    data = make_request(url_path, params)
    if data and data.get("analyses"):
        return data["analyses"][0].get("date")
    return None

def convert_to_numeric_or_na(value, as_type=None, decimal_places=None):
    """
    Converts a value to a specified numeric type (int or float) or returns "N/A".
    Applies decimal formatting if specified for floats.
    """
    if value is None or (isinstance(value, str) and str(value).strip().upper() == "N/A"):
        return "N/A"
    try:
        if as_type == int:
            return int(float(value))
        elif as_type == float:
            float_val = float(value)
            if decimal_places is not None:
                return round(float_val, decimal_places)
            return float_val
        else:
            return value
    except (ValueError, TypeError):
        return "N/A"

# ----------------------------- Excel/Data Structuring Functions -----------------------------

def create_headers(metric_keys):
    """
    Creates headers for the Excel sheet based on the metric keys.
    """
    headers = ["Project Name", "Project Key"]
    headers.extend(metric_keys.split(","))
    headers.extend([
        "Reliability Rating (A-E)",
        "Security Rating (A-E)",
        "Maintainability Rating (A-E)",
        "Security Hotspot Rating (A-E)",
        "last_analysis_date"
    ])
    return headers

def set_border(sheet, cell_range):
    """
    Sets a thin border around the specified cell range.
    """
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in sheet[cell_range]:
        for cell in row:
            cell.border = thin_border

def populate_sheet_with_data(sheet, projects_data, metric_keys_str):
    """
    Populates the Excel sheet with project data and metrics.
    """
    raw_metric_keys_list = metric_keys_str.split(',')
    excel_int_metrics = ['ncloc', 'bugs', 'reliability_rating', 'vulnerabilities',
                         'security_rating', 'security_review_rating', 'code_smells', 'sqale_rating']
    excel_float_metrics = ["duplicated_lines_density", "coverage"]


    for project in projects_data:
        row = [project["name"], project["key"]]
        metrics_map = {metric["metric"]: metric["value"] for metric in project["metrics"]}

        for metric_key in raw_metric_keys_list:
            value = metrics_map.get(metric_key)
            if metric_key in excel_int_metrics:
                row.append(convert_to_numeric_or_na(value, as_type=int))
            elif metric_key in excel_float_metrics:
                row.append(convert_to_numeric_or_na(value, as_type=float, decimal_places=1))
            else:
                row.append(value if value is not None else "N/A")

        row.extend([
            RATING_MAP.get(metrics_map.get("reliability_rating", "N/A"), "N/A"),
            RATING_MAP.get(metrics_map.get("security_rating", "N/A"), "N/A"),
            RATING_MAP.get(metrics_map.get("sqale_rating", "N/A"), "N/A"),
            RATING_MAP.get(metrics_map.get("security_review_rating", "N/A"), "N/A")
        ])
        row.append(project["last_analysis_date"])

        sheet.append(row)

def sort_and_clear_sheet(sheet):
    """
    Sorts the sheet data alphabetically by project name and clears any extra rows.
    """
    data = list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True))
    data.sort(key=lambda x: str(x[0]).lower())

    for row_idx in range(sheet.max_row, 1, -1):
        if row_idx > 1:
            sheet.delete_rows(row_idx)

    for row in data:
        sheet.append(list(row))


def auto_adjust_column_width(sheet):
    """
    Adjusts the column width in the sheet based on the longest value in each column.
    """
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    set_border(sheet, f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}")
    sheet.sheet_view.showGridLines = False

def align_headers(sheet, headers):
    """
    Aligns the headers in the sheet to the center.
    """
    for col in range(1, len(headers) + 1):
        sheet.cell(row=1, column=col).alignment = Alignment(horizontal="center")

def format_sheet(sheet, headers):
    """
    Applies formatting to the sheet.
    """
    create_table(sheet)
    freeze_top_row(sheet)
    auto_adjust_column_width(sheet)
    align_headers(sheet, headers)

def create_table(sheet):
    """
    Creates a table in the sheet for better formatting.
    """
    if sheet.max_row > 1:
        table = Table(displayName="SonarMetricsTable", ref=f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True
        )
        table.tableStyleInfo = style
        sheet.add_table(table)

def freeze_top_row(sheet):
    """
    Freezes the top row in the sheet.
    """
    sheet.freeze_panes = "B2"

def apply_summary_formatting(sheet):
    """
    Applies formatting to the summary sheet.
    """
    bold_font = Font(bold=True)
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    navy_blue_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center")

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = bold_font
            cell.fill = light_blue_fill

    for row_idx in [3, 5, 7, 12]:
        for cell in sheet[row_idx]:
            cell.font = white_font
            cell.fill = navy_blue_fill
            cell.alignment = center_alignment

    max_length = 0
    for cell in sheet['A']:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    sheet.column_dimensions['A'].width = max_length + 2

def apply_conditional_formatting_summary(sheet):
    """
    Applies conditional formatting to highlight specific cells in the summary sheet.
    """
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    dark_red_font = Font(color="9C0006", bold=True)

    rule = CellIsRule(
        operator="greaterThan",
        formula=["0"],
        fill=red_fill,
        font=dark_red_font
    )
    sheet.conditional_formatting.add("C4", rule)
    sheet.conditional_formatting.add("B6:E6", rule)
    sheet.conditional_formatting.add("C8:F11", rule)
    sheet.conditional_formatting.add("C13:F13", rule)


def highlight_na_rows_excel(sheet):
    """
    Highlights rows in light red where all values except 'Project Name' and 'Project Key' are 'N/A'.
    Checks from the 3rd column to the second-to-last column.
    """
    light_red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Changed to match HIGHLIGHT_NA_ROWS

    for row_idx in range(2, sheet.max_row + 1):
        row = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
        values_to_check = row[2:-1]
        if all(value == "N/A" for value in values_to_check):
            for col_idx in range(1, sheet.max_column + 1):
                sheet.cell(row=row_idx, column=col_idx).fill = light_red_fill


def generate_summary_data(metrics_sheet):
    """
    Generates aggregated metrics for the summary sheet.
    Returns a dictionary suitable for display.
    """
    summary_data = {}

    summary_data["Application Name"] = ORG_NAME
    total_repos = metrics_sheet.max_row - 1
    summary_data["Total Repos"] = total_repos

    alert_status_col_index = find_column_index(metrics_sheet, "alert_status")
    if alert_status_col_index is not None:
        quality_gate_counts = calculate_quality_gate_counts(metrics_sheet, alert_status_col_index)
        summary_data["Quality Gate - Passed"] = quality_gate_counts["OK"]
        summary_data["Quality Gate - Failed"] = quality_gate_counts["ERROR"]
        summary_data["Quality Gate - Not computed"] = quality_gate_counts["N/A"]

    coverage_col_index = find_column_index(metrics_sheet, "coverage")
    if coverage_col_index is not None:
        coverage_counts = calculate_coverage_counts(metrics_sheet, coverage_col_index)
        summary_data["Test Coverage - < 10%"] = coverage_counts["< 10%"]
        summary_data["Test Coverage - 10% - 30%"] = coverage_counts["10% - 30%"]
        summary_data["Test Coverage - 30% - 50%"] = coverage_counts["30% - 50%"]
        summary_data["Test Coverage - 50% - 80%"] = coverage_counts["50% - 80%"]
        summary_data["Test Coverage - > 80%"] = coverage_counts["> 80%"]
        summary_data["Test Coverage - N/A"] = coverage_counts["N/A"]

    reliability_col_index = find_column_index(metrics_sheet, "Reliability Rating (A-E)")
    if reliability_col_index is not None:
        reliability_counts = calculate_rating_counts(metrics_sheet, reliability_col_index)
        summary_data["Reliability Rating (Bugs) - A"] = reliability_counts["A"]
        summary_data["Reliability Rating (Bugs) - B"] = reliability_counts["B"]
        summary_data["Reliability Rating (Bugs) - C"] = reliability_counts["C"]
        summary_data["Reliability Rating (Bugs) - D"] = reliability_counts["D"]
        summary_data["Reliability Rating (Bugs) - E"] = reliability_counts["E"]
        summary_data["Reliability Rating (Bugs) - N/A"] = reliability_counts["N/A"]

    security_col_index = find_column_index(metrics_sheet, "Security Rating (A-E)")
    if security_col_index is not None:
        security_counts = calculate_rating_counts(metrics_sheet, security_col_index)
        summary_data["Security Rating (Vulnerabilities) - A"] = security_counts["A"]
        summary_data["Security Rating (Vulnerabilities) - B"] = security_counts["B"]
        summary_data["Security Rating (Vulnerabilities) - C"] = security_counts["C"]
        summary_data["Security Rating (Vulnerabilities) - D"] = security_counts["D"]
        summary_data["Security Rating (Vulnerabilities) - E"] = security_counts["E"]
        summary_data["Security Rating (Vulnerabilities) - N/A"] = security_counts["N/A"]

    maintainability_col_index = find_column_index(metrics_sheet, "Maintainability Rating (A-E)")
    if maintainability_col_index is not None:
        maintainability_counts = calculate_rating_counts(metrics_sheet, maintainability_col_index)
        summary_data["Maintainability Rating (Code Smells) - A"] = maintainability_counts["A"]
        summary_data["Maintainability Rating (Code Smells) - B"] = maintainability_counts["B"]
        summary_data["Maintainability Rating (Code Smells) - C"] = maintainability_counts["C"]
        summary_data["Maintainability Rating (Code Smells) - D"] = maintainability_counts["D"]
        summary_data["Maintainability Rating (Code Smells) - E"] = maintainability_counts["E"]
        summary_data["Maintainability Rating (Code Smells) - N/A"] = maintainability_counts["N/A"]

    hotspot_col_index = find_column_index(metrics_sheet, "Security Hotspot Rating (A-E)")
    if hotspot_col_index is not None:
        hotspot_counts = calculate_rating_counts(metrics_sheet, hotspot_col_index)
        summary_data["Security Hotspot Rating - A"] = hotspot_counts["A"]
        summary_data["Security Hotspot Rating - B"] = hotspot_counts["B"]
        summary_data["Security Hotspot Rating - C"] = hotspot_counts["C"]
        summary_data["Security Hotspot Rating - D"] = hotspot_counts["D"]
        summary_data["Security Hotspot Rating - E"] = hotspot_counts["E"]
        summary_data["Security Hotspot Rating - N/A"] = hotspot_counts["N/A"]

    duplication_col_index = find_column_index(metrics_sheet, "duplicated_lines_density")
    if duplication_col_index is not None:
        duplication_counts = calculate_duplication_counts(metrics_sheet, duplication_col_index)
        summary_data["Code Duplication - < 3%"] = duplication_counts["< 3%"]
        summary_data["Code Duplication - 3% - 5%"] = duplication_counts["3% - 5%"]
        summary_data["Code Duplication - 5% - 10%"] = duplication_counts["5% - 10%"]
        summary_data["Code Duplication - 10% - 20%"] = duplication_counts["10% - 20%"]
        summary_data["Code Duplication - > 20%"] = duplication_counts["> 20%"]
        summary_data["Code Duplication - N/A"] = duplication_counts["N/A"]

    return summary_data


def add_summary_sheet_to_workbook(workbook, metrics_sheet):
    """
    Adds a summary sheet to the workbook with aggregated metrics.
    """
    summary_sheet = workbook.create_sheet(title="Summary")

    summary_sheet["A1"] = "Application Name:"
    summary_sheet["B1"] = ORG_NAME

    total_repos = metrics_sheet.max_row - 1
    summary_sheet["A2"] = "Total Repos:"
    summary_sheet["B2"] = total_repos

    summary_sheet["A3"] = ""
    summary_sheet["B3"] = "Passed"
    summary_sheet["C3"] = "Failed"
    summary_sheet["D3"] = "Not computed"

    summary_sheet["A4"] = "Quality Gate"
    alert_status_col_index = find_column_index(metrics_sheet, "alert_status")
    if alert_status_col_index is not None:
        quality_gate_counts = calculate_quality_gate_counts(metrics_sheet, alert_status_col_index)
        summary_sheet["B4"] = quality_gate_counts["OK"]
        summary_sheet["C4"] = quality_gate_counts["ERROR"]
        summary_sheet["D4"] = quality_gate_counts["N/A"]

    summary_sheet["B5"] = "< 10%"
    summary_sheet["C5"] = "10% - 30%"
    summary_sheet["D5"] = "30% - 50%"
    summary_sheet["E5"] = "50% - 80%"
    summary_sheet["F5"] = "> 80%"
    summary_sheet["G5"] = "N/A"

    summary_sheet["A6"] = "Test Coverage"
    coverage_col_index = find_column_index(metrics_sheet, "coverage")
    if coverage_col_index is not None:
        coverage_counts = calculate_coverage_counts(metrics_sheet, coverage_col_index)
        summary_sheet["B6"] = coverage_counts["< 10%"]
        summary_sheet["C6"] = coverage_counts["10% - 30%"]
        summary_sheet["D6"] = coverage_counts["30% - 50%"]
        summary_sheet["E6"] = coverage_counts["50% - 80%"]
        summary_sheet["F6"] = coverage_counts["> 80%"]
        summary_sheet["G6"] = coverage_counts["N/A"]

    summary_sheet["A7"] = ""
    summary_sheet["B7"] = "A"
    summary_sheet["C7"] = "B"
    summary_sheet["D7"] = "C"
    summary_sheet["E7"] = "D"
    summary_sheet["F7"] = "E"
    summary_sheet["G7"] = "N/A"

    summary_sheet["A8"] = "Reliability Rating (Bugs)"
    reliability_col_index = find_column_index(metrics_sheet, "Reliability Rating (A-E)")
    if reliability_col_index is not None:
        reliability_counts = calculate_rating_counts(metrics_sheet, reliability_col_index)
        summary_sheet["B8"] = reliability_counts["A"]
        summary_sheet["C8"] = reliability_counts["B"]
        summary_sheet["D8"] = reliability_counts["C"]
        summary_sheet["E8"] = reliability_counts["D"]
        summary_sheet["F8"] = reliability_counts["E"]
        summary_sheet["G8"] = reliability_counts["N/A"]

    summary_sheet["A9"] = "Security Rating (Vulnerabilities)"
    security_col_index = find_column_index(metrics_sheet, "Security Rating (A-E)")
    if security_col_index is not None:
        security_counts = calculate_rating_counts(metrics_sheet, security_col_index)
        summary_sheet["B9"] = security_counts["A"]
        summary_sheet["C9"] = security_counts["B"]
        summary_sheet["D9"] = security_counts["C"]
        summary_sheet["E9"] = security_counts["D"]
        summary_sheet["F9"] = security_counts["E"]
        summary_sheet["G9"] = security_counts["N/A"]

    summary_sheet["A10"] = "Maintainability Rating (Code Smells)"
    maintainability_col_index = find_column_index(metrics_sheet, "Maintainability Rating (A-E)")
    if maintainability_col_index is not None:
        maintainability_counts = calculate_rating_counts(metrics_sheet, maintainability_col_index)
        summary_sheet["B10"] = maintainability_counts["A"]
        summary_sheet["C10"] = maintainability_counts["B"]
        summary_sheet["D10"] = maintainability_counts["C"]
        summary_sheet["E10"] = maintainability_counts["D"]
        summary_sheet["F10"] = maintainability_counts["E"]
        summary_sheet["G10"] = maintainability_counts["N/A"]

    summary_sheet["A11"] = "Security Hotspot Rating"
    hotspot_col_index = find_column_index(metrics_sheet, "Security Hotspot Rating (A-E)")
    if hotspot_col_index is not None:
        hotspot_counts = calculate_rating_counts(metrics_sheet, hotspot_col_index)
        summary_sheet["B11"] = hotspot_counts["A"]
        summary_sheet["C11"] = hotspot_counts["B"]
        summary_sheet["D11"] = hotspot_counts["C"]
        summary_sheet["E11"] = hotspot_counts["D"]
        summary_sheet["F11"] = hotspot_counts["E"]
        summary_sheet["G11"] = hotspot_counts["N/A"]

    summary_sheet["A12"] = ""
    summary_sheet["B12"] = "< 3%"
    summary_sheet["C12"] = "3% - 5%"
    summary_sheet["D12"] = "5% - 10%"
    summary_sheet["E12"] = "10% - 20%"
    summary_sheet["F12"] = "> 20%"
    summary_sheet["G12"] = "N/A"

    summary_sheet["A13"] = "Code Duplication"
    duplication_col_index = find_column_index(metrics_sheet, "duplicated_lines_density")
    if duplication_col_index is not None:
        duplication_counts = calculate_duplication_counts(metrics_sheet, duplication_col_index)
        summary_sheet["B13"] = duplication_counts["< 3%"]
        summary_sheet["C13"] = duplication_counts["3% - 5%"]
        summary_sheet["D13"] = duplication_counts["5% - 10%"]
        summary_sheet["E13"] = duplication_counts["10% - 20%"]
        summary_sheet["F13"] = duplication_counts["> 20%"]
        summary_sheet["G13"] = duplication_counts["N/A"]

    apply_summary_formatting(summary_sheet)
    apply_conditional_formatting_summary(summary_sheet)
    auto_adjust_column_width(summary_sheet)


# ----------------------------- Calculation Functions -----------------------------\

def calculate_duplication_counts(sheet, col_index):
    """
    Calculates the counts for code duplication (< 3%, 3% - 5%, 5% - 10%, 10% - 20%, > 20%,  N/A).
    """
    counts = {"< 3%": 0, "3% - 5%": 0, "5% - 10%": 0, "10% - 20%": 0, "> 20%": 0, "N/A": 0}
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_index, max_col=col_index):
        value = row[0].value
        if isinstance(value, (int, float)):
            if value < 3:
                counts["< 3%"] += 1
            elif 3 <= value < 5:
                counts["3% - 5%"] += 1
            elif 5 <= value < 10:
                counts["5% - 10%"] += 1
            elif 10 <= value < 20:
                counts["10% - 20%"] += 1
            elif value >= 20:
                counts["> 20%"] += 1
        else:
            counts["N/A"] += 1
    return counts

def calculate_rating_counts(sheet, col_index):
    """
    Calculates the counts for ratings (A, B, C, D, E, N/A).
    """
    counts = {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "N/A": 0}
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_index, max_col=col_index):
        value = row[0].value
        if value in counts:
            counts[value] += 1
    return counts

def calculate_coverage_counts(sheet, col_index):
    """
    Calculates the counts for code coverage (< 10%, 10% - 30%, 30% - 50%, 50% - 80%, > 80%,  N/A).
    """
    counts = {"< 10%": 0, "10% - 30%": 0, "30% - 50%": 0, "50% - 80%": 0, "> 80%": 0, "N/A": 0}
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_index, max_col=col_index):
        value = row[0].value
        if isinstance(value, (int, float)):
            if value < 10:
                counts["< 10%"] += 1
            elif 10 <= value < 30:
                counts["10% - 30%"] += 1
            elif 30 <= value < 50:
                counts["30% - 50%"] += 1
            elif 50 <= value < 80:
                counts["50% - 80%"] += 1
            elif value >= 80:
                counts["> 80%"] += 1
        else:
            counts["N/A"] += 1
    return counts

def calculate_quality_gate_counts(sheet, col_index):
    """
    Calculates the counts for Quality Gate statuses (Passed, Failed, Not computed).
    """
    counts = {"OK": 0, "ERROR": 0, "N/A": 0}
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_index, max_col=col_index):
        value = row[0].value
        if value == "OK":
            counts["OK"] += 1
        elif value == "ERROR":
            counts["ERROR"] += 1
        else:
            counts["N/A"] += 1
    return counts

def find_column_index(sheet, column_name):
    """
    Finds the index of a column by its name.
    """
    for col_idx, cell in enumerate(sheet[1], start=1):
        if cell.value == column_name:
            return col_idx
    return None

def process_project(project, metric_keys):
    """
    Processes a single project to fetch its metrics.
    """
    project_key = project.get("key")
    project_name = project.get("name")
    metrics = fetch_project_metrics(project_key, metric_keys)

    last_analysis_date = fetch_last_analysis_date(project_key)

    if not metrics:
        return {"name": project_name, "key": project_key, "metrics": [], "last_analysis_date": last_analysis_date}

    return {"name": project_name, "key": project_key, "metrics": metrics, "last_analysis_date": last_analysis_date}

def generate_excel_file(projects_data, metric_keys):
    """
    Generates the Excel file (in-memory) with raw metrics and summary.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sonar Metrics"

    headers = create_headers(metric_keys)
    sheet.append(headers)

    populate_sheet_with_data(sheet, projects_data, metric_keys)
    # st.success("Populated sheet with data for Excel export.") # Changed to st.success

    sort_and_clear_sheet(sheet)
    # st.success("Sorted and cleared the metrics sheet.") # Changed to st.success

    format_sheet(sheet, headers)
    # st.success("Formatted the metrics sheet.") # Changed to st.success

    highlight_na_rows_excel(sheet)
    # st.success("Highlighted rows with all 'N/A' values in metrics sheet.") # Changed to st.success

    auto_adjust_column_width(sheet)

    add_summary_sheet_to_workbook(workbook, sheet)

    excel_buffer = io.BytesIO()
    try:
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer
    except Exception as e:
        st.error(f"Error saving workbook: {e}")
        return None

# --- UI Specific Styling Functions ---

def highlight_na_rows_dataframe(row):
    """
    Applies light red background to rows where all values (excluding Project Name/Key) are 'N/A'.
    This function should be applied row-wise.
    """
    cols_to_check = [
        'alert_status', 'ncloc', 'bugs', 'reliability_rating',
        'vulnerabilities', 'security_rating', 'security_review_rating',
        'code_smells', 'sqale_rating', 'duplicated_lines_density', 'coverage',
        'Reliability Rating (A-E)', 'Security Rating (A-E)',
        'Maintainability Rating (A-E)', 'Security Hotspot Rating (A-E)'
    ]

    existing_cols_to_check = [col for col in cols_to_check if col in row.index]

    if not existing_cols_to_check:
        return [''] * len(row)

    is_na_row = True
    for col in existing_cols_to_check:
        val = row[col]
        if pd.isna(val) or val is None or (isinstance(val, str) and val.strip().upper() != 'N/A'):
            is_na_row = False
            break

    if is_na_row:
        return [HIGHLIGHT_NA_ROWS] * len(row)
    return [''] * len(row)

# Custom styling for Quality Gate counts
def apply_quality_gate_styles(col_data):
    styles = [''] * len(col_data)
    for i, count in enumerate(col_data):
        if col_data.index[i] == 'Failed' and count > 0:
            styles[i] = DATA_NEED_ATTENTION
    return styles

# Custom styling for Test Coverage counts
def apply_coverage_styles(col_data):
    styles = [''] * len(col_data)
    for i, count in enumerate(col_data):
        current_range = col_data.index[i]
        # Highlight if count > 0 AND range is problematic (< 80%) AND NOT N/A
        if current_range != "> 80%" and current_range.upper() != "N/A" and isinstance(count, (int, float)) and count > 0:
             styles[i] = DATA_NEED_ATTENTION
    return styles

# Custom styling for Code Duplication counts
def apply_duplication_styles(col_data):
    styles = [''] * len(col_data)
    for i, count in enumerate(col_data):
        current_range = col_data.index[i]
        # Highlight if count > 0 AND range is problematic (> 0%) AND NOT N/A
        # The '< 3%' range is generally considered good, so exclude it from problematic.
        if current_range != "< 3%" and current_range.upper() != "N/A" and isinstance(count, (int, float)) and count > 0:
             styles[i] = DATA_NEED_ATTENTION
    return styles

# Custom styling for Rating Distributions
def apply_ratings_styles(col_data):
    styles = [''] * len(col_data)
    for i, count in enumerate(col_data):
        current_rating = col_data.index[i]
        # Highlight if rating is B, C, D, E and count > 0
        if current_rating in ["B", "C", "D", "E"] and isinstance(count, (int, float)) and count > 0:
            styles[i] = DATA_NEED_ATTENTION
    return styles


# ----------------------------- Main Streamlit App -----------------------------

def main_streamlit():
    global auth, SONAR_BASE_URL, ORG_NAME, ORG_KEY

    # Sidebar for inputs (Sonar URL, Token, and Org Key)
    with st.sidebar:
        st.header("SonarCloud Configuration")
        # st.markdown("Enter your SonarCloud credentials and URL to fetch project metrics.")

        sonar_url_input = st.text_input("SonarCloud URL", value="https://sonarcloud.io", help="The base URL of your SonarCloud instance (e.g., 'https://sonarcloud.io').")
        sonar_token_input = st.text_input("Sonar Token", type="password", help="Your SonarCloud personal access token.")
        organization_name_input = st.text_input("Organization Name", value="<org_name>", help="Your SonarCloud organization Name (e.g., 'XYZ').")
        organization_key_input = st.text_input("Organization Key", value="<org_key>", help="Your SonarCloud organization key (e.g., 'xyzinc').")

        fetch_button = st.button("Fetch SonarCloud Metrics")

        if fetch_button:
            if not sonar_url_input:
                st.error("SonarCloud URL not found. Please enter the URL.")
                return
            if not sonar_token_input:
                st.error("Sonar Token not found. Please enter your 'SONAR_TOKEN'.")
                return
            if not organization_name_input:
                st.error("Organization Name not found. Please enter your 'ORGANIZATION_NAME'.")
                return
            if not organization_key_input:
                st.error("Organization Key not found. Please enter your 'ORGANIZATION_KEY'.")
                return
            
            # Set global variables from UI input
            SONAR_BASE_URL = sonar_url_input.rstrip('/')
            ORG_NAME = organization_name_input.rstrip('/')
            ORG_KEY = organization_key_input.rstrip('/')
            auth = (sonar_token_input, "")

            metric_keys_str = (
                "alert_status,ncloc,"
                "bugs,reliability_rating,"
                "vulnerabilities,security_rating,"
                "security_review_rating,code_smells,sqale_rating,"
                "duplicated_lines_density,coverage"
            )
            raw_metric_keys_list = metric_keys_str.split(',')

            projects_data_for_ui_raw = []
            raw_projects_data_for_excel = []
            start_time = datetime.now()

            with st.spinner("Connecting to SonarCloud and fetching project list..."):
                projects = fetch_projects(ORG_KEY)
                if not projects:
                    st.warning("No projects found for the given organization key. Please check your Organization Key and Token.")
                    st.session_state['data_fetched'] = False
                    return
                st.success(f"Successfully connected! Found {len(projects)} projects.")

            progress_text = "Processing project metrics..."
            my_bar = st.progress(0, text=progress_text)
            processed_count = 0

            ui_numeric_metrics = {
                'ncloc': int,
                'bugs': int,
                'reliability_rating': int,
                'vulnerabilities': int,
                'security_rating': int,
                'security_review_rating': int,
                'code_smells': int,
                'sqale_rating': int,
                'duplicated_lines_density': float,
                'coverage': float,
            }

            with ThreadPoolExecutor() as executor:
                future_to_project = {executor.submit(process_project, project, metric_keys_str): project for project in projects}
                for future in as_completed(future_to_project):
                    try:
                        project_data = future.result()
                        raw_projects_data_for_excel.append(project_data)

                        ui_row = {
                            "Project Name": project_data["name"],
                            "Project Key": project_data["key"]
                        }
                        metrics_map = {metric["metric"]: metric["value"] for metric in project_data["metrics"]}

                        for metric_key in raw_metric_keys_list:
                            value = metrics_map.get(metric_key)
                            target_type = ui_numeric_metrics.get(metric_key)
                            if target_type == float:
                                ui_row[metric_key] = convert_to_numeric_or_na(value, as_type=float, decimal_places=1)
                            elif target_type == int:
                                ui_row[metric_key] = convert_to_numeric_or_na(value, as_type=int)
                            else:
                                ui_row[metric_key] = value if value is not None else "N/A"

                        ui_row["Reliability Rating (A-E)"] = RATING_MAP.get(metrics_map.get("reliability_rating", "N/A"), "N/A")
                        ui_row["Security Rating (A-E)"] = RATING_MAP.get(metrics_map.get("security_rating", "N/A"), "N/A")
                        ui_row["Maintainability Rating (A-E)"] = RATING_MAP.get(metrics_map.get("sqale_rating", "N/A"), "N/A")
                        ui_row["Security Hotspot Rating (A-E)"] = RATING_MAP.get(metrics_map.get("security_review_rating", "N/A"), "N/A")
                        
                        ui_row["last_analysis_date"] = project_data["last_analysis_date"] if project_data["last_analysis_date"] is not None else "N/A"

                        projects_data_for_ui_raw.append(ui_row)

                        processed_count += 1
                        my_bar.progress(processed_count / len(projects), text=progress_text)
                    except Exception as e:
                        st.error(f"Error processing project: {future_to_project[future].get('name', 'Unknown')} - {e}")
            my_bar.empty()

            st.session_state['projects_data_for_ui'] = projects_data_for_ui_raw
            st.session_state['raw_projects_data_for_excel'] = raw_projects_data_for_excel
            st.session_state['metric_keys'] = metric_keys_str
            st.session_state['data_fetched'] = True

            end_time = datetime.now()
            st.success(f"Data fetching complete! Duration: {end_time - start_time}")

    # Main content area (right of the sidebar)
    # st.header("SonarCloud Metrics Overview")

    if 'data_fetched' in st.session_state and st.session_state['data_fetched']:
        # Section 1: Raw Project Metrics
        with st.expander("View Raw Project Metrics", expanded=True):
            if st.session_state['projects_data_for_ui']:
                df_metrics = pd.DataFrame(st.session_state['projects_data_for_ui'])
                
                display_columns_order = [
                    'Project Name', 'Project Key', 'alert_status', 'coverage',
                    'ncloc', 'bugs', 'reliability_rating', 'vulnerabilities', 
                    'security_rating', 'security_review_rating', 'code_smells', 'sqale_rating',
                    'duplicated_lines_density', 'Reliability Rating (A-E)', 'Security Rating (A-E)',
                    'Maintainability Rating (A-E)', 'Security Hotspot Rating (A-E)',
                    'last_analysis_date'
                ]

                for col in display_columns_order:
                    if col not in df_metrics.columns:
                        df_metrics[col] = "N/A"
                
                df_metrics = df_metrics[display_columns_order]

                # Apply string formatting for 'duplicated_lines_density' and 'coverage'
                for col_name in ['duplicated_lines_density', 'coverage']:
                    if col_name in df_metrics.columns:
                        # Ensure 'N/A' values are not attempted to be formatted as floats
                        df_metrics[col_name] = df_metrics[col_name].apply(
                            lambda x: f"{x:.1f}" if isinstance(x, (float, int)) and not np.isnan(x) else x
                        )

                # Sort by Project Name
                # Clean 'Project Name' for consistent sorting (strip whitespace and convert to lower case)
                df_metrics['Project Name_sort_key'] = df_metrics['Project Name'].astype(str).str.strip().str.lower()

                # Sort by the new sort key column
                df_metrics_sorted = df_metrics.sort_values(by="Project Name_sort_key").reset_index(drop=True)

                # Drop the temporary sort key column before displaying
                df_metrics_sorted = df_metrics_sorted.drop(columns=['Project Name_sort_key'])

                # Start index from 1 for better readability
                df_metrics_sorted.index = np.arange(1, len(df_metrics_sorted) + 1)

                st.dataframe(df_metrics_sorted.style.apply(highlight_na_rows_dataframe, axis=1), use_container_width=True)
            else:
                st.info("No project metrics available to display.")

        # Create a temporary workbook to generate summary data for UI display
        temp_workbook_for_summary = Workbook()
        temp_sheet_for_summary = temp_workbook_for_summary.active
        temp_sheet_for_summary.title = "Temp Sonar Metrics"
        headers_for_temp = create_headers(st.session_state['metric_keys'])
        temp_sheet_for_summary.append(headers_for_temp)
        populate_sheet_with_data(temp_sheet_for_summary, st.session_state['raw_projects_data_for_excel'], st.session_state['metric_keys'])

        # Section 2: Aggregated Summary
        with st.expander("View Aggregated Summary", expanded=True):
            summary_data = generate_summary_data(temp_sheet_for_summary)
            st.session_state['summary_data'] = summary_data

            # st.subheader("Application Overview")
            st.markdown(f"""
                - **Application Name:** `{summary_data.get('Application Name')}`
                - **Total Repositories:** `{summary_data.get('Total Repos')}`
            """)

            col_qg, col_tc, col_cd, col_ratings = st.columns([1, 1, 1, 3])
            
            # For Quality Gate, Test Coverage, Code Duplication and Rating
            with col_qg:
                # st.subheader("Quality Gate Status")
                st.markdown(
                    "<h4 style='font-size:18px;'>Quality Gate Status</h4>", 
                    unsafe_allow_html=True
                )
                quality_gate_df = pd.DataFrame({
                    "Status": ["Passed", "Failed", "Not Computed"],
                    "Count": [
                        summary_data.get("Quality Gate - Passed", 0),
                        summary_data.get("Quality Gate - Failed", 0),
                        summary_data.get("Quality Gate - Not computed", 0)
                    ]
                })
                quality_gate_df_indexed = quality_gate_df.set_index("Status")
                styled_quality_gate_df = quality_gate_df_indexed.style.apply(
                    apply_quality_gate_styles, axis=0, subset=['Count']
                )
                st.dataframe(styled_quality_gate_df, use_container_width=True)

            # Row 2 for Test Coverage and Code Duplication (2 columns)
            with col_tc:
                # st.subheader("Test Coverage Distribution")
                st.markdown(
                    "<h4 style='font-size:18px;'>Test Coverage Distribution</h4>", 
                    unsafe_allow_html=True
                )
                coverage_df = pd.DataFrame({
                    "Range": ["< 10%", "10% - 30%", "30% - 50%", "50% - 80%", "> 80%", "N/A"],
                    "Count": [
                        summary_data.get("Test Coverage - < 10%", 0),
                        summary_data.get("Test Coverage - 10% - 30%", 0),
                        summary_data.get("Test Coverage - 30% - 50%", 0),
                        summary_data.get("Test Coverage - 50% - 80%", 0),
                        summary_data.get("Test Coverage - > 80%", 0),
                        summary_data.get("Test Coverage - N/A", 0)
                    ]
                })
                coverage_df_indexed = coverage_df.set_index("Range")
                styled_coverage_df = coverage_df_indexed.style.apply(
                    apply_coverage_styles, axis=0, subset=['Count']
                )
                st.dataframe(styled_coverage_df, use_container_width=True)

            with col_cd:
                # st.subheader("Code Duplication Distribution")
                st.markdown(
                    "<h4 style='font-size:18px;'>Code Duplication Distribution</h4>", 
                    unsafe_allow_html=True
                )
                duplication_df = pd.DataFrame({
                    "Range": ["< 3%", "3% - 5%", "5% - 10%", "10% - 20%", "> 20%", "N/A"],
                    "Count": [
                        summary_data.get("Code Duplication - < 3%", 0),
                        summary_data.get("Code Duplication - 3% - 5%", 0),
                        summary_data.get("Code Duplication - 5% - 10%", 0),
                        summary_data.get("Code Duplication - 10% - 20%", 0),
                        summary_data.get("Code Duplication - > 20%", 0),
                        summary_data.get("Code Duplication - N/A", 0)
                    ]
                })
                duplication_df_indexed = duplication_df.set_index("Range")
                styled_duplication_df = duplication_df_indexed.style.apply(
                    apply_duplication_styles, axis=0, subset=['Count']
                )
                st.dataframe(styled_duplication_df, use_container_width=True)

            with col_ratings:
                # st.subheader("Rating Distributions")
                st.markdown(
                    "<h4 style='font-size:18px;'>Rating Distributions</h4>", 
                    unsafe_allow_html=True
                )
                ratings_df = pd.DataFrame({
                    "Rating": ["A", "B", "C", "D", "E", "N/A"],
                    "Reliability (Bugs)": [
                        summary_data.get("Reliability Rating (Bugs) - A", 0),
                        summary_data.get("Reliability Rating (Bugs) - B", 0),
                        summary_data.get("Reliability Rating (Bugs) - C", 0),
                        summary_data.get("Reliability Rating (Bugs) - D", 0),
                        summary_data.get("Reliability Rating (Bugs) - E", 0),
                        summary_data.get("Reliability Rating (Bugs) - N/A", 0)
                    ],
                    "Security (Vulnerabilities)": [
                        summary_data.get("Security Rating (Vulnerabilities) - A", 0),
                        summary_data.get("Security Rating (Vulnerabilities) - B", 0),
                        summary_data.get("Security Rating (Vulnerabilities) - C", 0),
                        summary_data.get("Security Rating (Vulnerabilities) - D", 0),
                        summary_data.get("Security Rating (Vulnerabilities) - E", 0),
                        summary_data.get("Security Rating (Vulnerabilities) - N/A", 0)
                    ],
                    "Maintainability (Code Smells)": [
                        summary_data.get("Maintainability Rating (Code Smells) - A", 0),
                        summary_data.get("Maintainability Rating (Code Smells) - B", 0),
                        summary_data.get("Maintainability Rating (Code Smells) - C", 0),
                        summary_data.get("Maintainability Rating (Code Smells) - D", 0),
                        summary_data.get("Maintainability Rating (Code Smells) - E", 0),
                        summary_data.get("Maintainability Rating (Code Smells) - N/A", 0)
                    ],
                    "Security Hotspot": [
                        summary_data.get("Security Hotspot Rating - A", 0),
                        summary_data.get("Security Hotspot Rating - B", 0),
                        summary_data.get("Security Hotspot Rating - C", 0),
                        summary_data.get("Security Hotspot Rating - D", 0),
                        summary_data.get("Security Hotspot Rating - E", 0),
                        summary_data.get("Security Hotspot Rating - N/A", 0)
                    ]
                })
                ratings_df_indexed = ratings_df.set_index("Rating")
                styled_ratings_df = ratings_df_indexed.style.apply(
                    apply_ratings_styles, axis=0, subset=["Reliability (Bugs)", "Security (Vulnerabilities)", "Maintainability (Code Smells)", "Security Hotspot"]
                )
                st.dataframe(styled_ratings_df, use_container_width=True)


        # Only show download button if data is fetched and available
        if st.session_state['raw_projects_data_for_excel']:
            excel_file_buffer = generate_excel_file(st.session_state['raw_projects_data_for_excel'], st.session_state['metric_keys'])
            if excel_file_buffer:
                st.download_button(
                    label="Download SonarMetrics.xlsx",
                    data=excel_file_buffer,
                    file_name="SonarMetrics.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.info("Enter your SonarCloud credentials in the sidebar and click 'Fetch SonarCloud Metrics' to see the data.")

if __name__ == "__main__":
    if 'data_fetched' not in st.session_state:
        st.session_state['data_fetched'] = False
    if 'projects_data_for_ui' not in st.session_state:
        st.session_state['projects_data_for_ui'] = []
    if 'raw_projects_data_for_excel' not in st.session_state:
        st.session_state['raw_projects_data_for_excel'] = []
    if 'metric_keys' not in st.session_state:
        st.session_state['metric_keys'] = ""
    if 'summary_data' not in st.session_state:
        st.session_state['summary_data'] = {}

    main_streamlit()